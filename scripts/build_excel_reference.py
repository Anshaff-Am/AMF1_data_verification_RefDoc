#!/usr/bin/env python3
"""
AMF1 Phase 0 — Excel Reference Document Builder

Reads reference_document.csv and produces a human-readable Excel workbook
with one sheet per dashboard section. Only extracts key dashboard metrics
(Total sample + country/demographic filter slices from Banner1).

Output: reference_document_REVIEW.xlsx
"""

import csv
import sys
import io
from pathlib import Path
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

BASE_DIR  = Path(__file__).parent.parent
INPUT_CSV = BASE_DIR / "outputs" / "reference_document.csv"
OUTPUT_XL = BASE_DIR / "outputs" / "reference_document_REVIEW.xlsx"

# ── Filter columns to include (from Banner1, in display order) ─────────────
# Maps filter_option value → display header label
FILTER_COLS_B1 = {
    "Total":        "Total",
    # Countries
    "Australia":    "Australia",
    "China":        "China",
    "Japan":        "Japan",
    "UK":           "UK",
    "USA":          "USA",
    "Saudi Arabia": "Saudi Arabia",
    "Mexico":       "Mexico",
    # Gender
    "Men":          "Men",
    "Women":        "Women",
    # Age
    "18-34":        "18-34",
    "35-54":        "35-54",
    "55+":          "55+",
    # Sports / F1
    "Sports Fan":   "Sports Fan",
    "Non-Sports Fan": "Non-Sports Fan",
    # BDMs
    "Any":          "BDMs (Any)",
}

# ── Which response rows to keep for each question section ─────────────────
# Maps: (wave, question_code, substring_in_question_text) → row_label_substrings_to_keep
# Empty list = keep all response rows

KEY_METRICS = [
    # Overview page P0
    ("D1",  "",                  []),                          # all brands
    ("D2",  "Top 2 Box",        []),                          # all brands per row
    ("D3",  "",                  ["Currently", "Consider"]),  # 2 rows per brand
    ("D4",  "Top 2 Box",        []),                          # all brands
    ("D5",  "NPS Score",        []),                          # all brands
    ("D6",  "Total Sample",     []),                          # all brands
    # Honda
    ("C5",  "",                  []),
    ("C6",  "Top 2 Box",        []),
    # CoreWeave
    ("D7A", "",                  []),
    ("D8",  "",                  []),
    ("D9",  "Top 2 Box",        []),
    ("COR1","Top 2 Box",        []),
    # Aramco
    ("ARC1","Top 2 Box",        []),
    ("ARC2","Top 2 Box",        []),
    ("ARC5","",                  []),
    ("ARC6","",                  []),
    ("ARC7","",                  []),
    ("ARC8","",                  []),
    # Coinbase
    ("CON1","",                  []),
    ("CON2","",                  []),
    ("CON3","",                  []),
    ("CON4","",                  []),
    ("CON5","",                  []),
    ("CON6","Top 2 Box",        []),
    ("CON7","",                  []),
    ("CON8","Top 2 Box",        []),
    # Maaden
    ("MAA1","",                  []),
    ("MAA2","",                  []),
    ("MAA3","",                  []),
    ("MAA4","",                  []),
    ("MAA5","",                  []),
    ("MAA6","",                  []),
    ("MAA7","Top 2 Box",        []),
    ("MAA8","",                  []),
    # Valvoline
    ("VAL1","Top 2 Box",        []),
    # Pepperstone
    ("PEP1","",                  []),
    ("PEP2","",                  []),
    # NexGen
    ("NEX1","",                  []),
    # Glenfiddich
    ("GLN1","Top 2 Box",        []),
    # ServiceNow
    ("SER1","",                  []),
    # Atlas Air
    ("ATL1","",                  []),
    ("ATL2","Top 2 Box",        []),
    # Cognizant
    ("COG1","Top 2 Box",        []),
    ("COG2","",                  []),
    ("COG3","",                  []),
    # TikTok
    ("TTK1","",                  []),
    ("TTK2","",                  []),
    ("TTK3","",                  []),
    ("TTK4","",                  []),
    ("TTK5","",                  []),
    ("TTK6","",                  []),
    ("TTK7","",                  []),
    # Aston Martin Lagonda
    ("B1",  "",                  []),
    ("B2",  "",                  []),
    ("B2A", "",                  []),
    ("B3",  "NPS Score",        []),
    ("B5",  "",                  []),
    ("B7",  "Top 2 Box",        []),
    # D14/15/16 partnership
    ("D14", "Top 2 Box",        []),
    ("D15", "",                  []),
    ("D16", "",                  []),
]

# Dashboard section groupings → sheet names
SECTIONS = {
    "1_Overview_D1-D6":   ["D1","D2","D3","D4","D5","D6"],
    "2_Honda_C5-C6":      ["C5","C6"],
    "3_CoreWeave":        ["D7A","D8","D9","COR1"],
    "4_Aramco":           ["ARC1","ARC2","ARC5","ARC6","ARC7","ARC8"],
    "5_Coinbase":         ["CON1","CON2","CON3","CON4","CON5","CON6","CON7","CON8"],
    "6_Maaden":           ["MAA1","MAA2","MAA3","MAA4","MAA5","MAA6","MAA7","MAA8"],
    "7_Valvoline":        ["VAL1"],
    "8_Pepperstone":      ["PEP1","PEP2"],
    "9_NexGen":           ["NEX1"],
    "10_Glenfiddich":     ["GLN1"],
    "11_ServiceNow":      ["SER1"],
    "12_AtlasAir":        ["ATL1","ATL2"],
    "13_Cognizant":       ["COG1","COG2","COG3"],
    "14_TikTok":          ["TTK1","TTK2","TTK3","TTK4","TTK5","TTK6","TTK7"],
    "15_AstonMartinLag":  ["B1","B2","B2A","B3","B5","B7"],
    "16_Partnership":     ["D14","D15","D16"],
}

# ── Excel style helpers ────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")    # dark navy
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
SUBHDR_FILL= PatternFill("solid", fgColor="2E75B6")    # mid blue
SUBHDR_FONT= Font(bold=True, color="FFFFFF", size=10)
LABEL_FILL = PatternFill("solid", fgColor="D9E2F3")    # light blue
W1_FILL    = PatternFill("solid", fgColor="E2EFDA")    # light green for W1
W2_FILL    = PatternFill("solid", fgColor="FFF2CC")    # light yellow for W2
THIN       = Side(style="thin", color="BFBFBF")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",  vertical="center", wrap_text=True)


def style_cell(cell, font=None, fill=None, align=None, border=None):
    if font:   cell.font      = font
    if fill:   cell.fill      = fill
    if align:  cell.alignment = align
    if border: cell.border    = border


def build_lookup(rows):
    """
    Build nested dict:
    lookup[wave][qcode][question_text][response_label][filter_option] = value
    Using Banner1 only (primary source for country/demo slices).
    """
    lookup = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: defaultdict(dict))))

    for r in rows:
        if r["banner"] != "Banner1":
            continue
        fo = r["filter_option"]
        if fo not in FILTER_COLS_B1:
            continue
        key = (r["wave"], r["question_code"], r["question_text"], r["response_label"])
        if r["value"]:
            lookup[r["wave"]][r["question_code"]][r["question_text"]][r["response_label"]][fo] = r["value"]

    return lookup


def write_section_sheet(ws, section_qcodes, lookup, waves=("W1", "W2")):
    """Write a section to a worksheet."""
    filter_keys = list(FILTER_COLS_B1.keys())
    filter_hdrs = list(FILTER_COLS_B1.values())

    # Build metric filter: qcode → (text_substring, row_label_substrings)
    metric_map = {}
    for (qc, txt_sub, row_subs) in KEY_METRICS:
        metric_map[qc] = (txt_sub, row_subs)

    col_offset = 4  # cols A-C = Wave / Question / Response label

    # Header row 1: Wave labels
    # Header row 2: Filter option labels
    header1 = ws.row_dimensions[1]
    header1.height = 30

    # Write column headers
    ws.cell(1, 1, "Wave")
    ws.cell(1, 2, "Question Code")
    ws.cell(1, 3, "Question Text")
    ws.cell(1, 4, "Response / Brand")
    for j, lbl in enumerate(filter_hdrs, start=col_offset + 1):
        ws.cell(1, j, lbl)

    for c in range(1, col_offset + len(filter_hdrs) + 1):
        cell = ws.cell(1, c)
        style_cell(cell, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)

    current_row = 2

    for wave in waves:
        if wave not in lookup:
            continue
        wave_data = lookup[wave]

        for qcode in section_qcodes:
            if qcode not in wave_data:
                continue

            txt_sub, row_subs = metric_map.get(qcode, ("", []))

            for q_text, resp_map in sorted(wave_data[qcode].items()):
                # Filter by question_text substring
                if txt_sub and txt_sub not in q_text:
                    continue

                for resp_label, filter_vals in sorted(resp_map.items()):
                    # Filter by response row substring
                    if row_subs:
                        if not any(s.lower() in resp_label.lower() for s in row_subs):
                            continue

                    # Write data row
                    fill = W1_FILL if wave == "W1" else W2_FILL

                    ws.cell(current_row, 1, wave)
                    ws.cell(current_row, 2, qcode)
                    # Truncate question_text for readability
                    short_q = q_text[:80] + ("…" if len(q_text) > 80 else "")
                    ws.cell(current_row, 3, short_q)
                    ws.cell(current_row, 4, resp_label)

                    for j, fk in enumerate(filter_keys, start=col_offset + 1):
                        val = filter_vals.get(fk, "")
                        if val:
                            try:
                                ws.cell(current_row, j, float(val))
                            except (ValueError, TypeError):
                                ws.cell(current_row, j, val)
                        else:
                            ws.cell(current_row, j, "")

                    # Style row
                    for c in range(1, col_offset + len(filter_hdrs) + 1):
                        cell = ws.cell(current_row, c)
                        style_cell(cell, border=BORDER)
                        if c <= 4:
                            style_cell(cell, fill=LABEL_FILL if c == 4 else fill, align=LEFT)
                        else:
                            style_cell(cell, align=CENTER)
                        if c in (1, 2):
                            style_cell(cell, fill=fill)

                    current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 8   # Wave
    ws.column_dimensions["B"].width = 12  # Q code
    ws.column_dimensions["C"].width = 55  # Q text
    ws.column_dimensions["D"].width = 40  # Response
    for j in range(col_offset + 1, col_offset + len(filter_hdrs) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13

    # Freeze first row + first 4 cols
    ws.freeze_panes = "E2"

    return current_row - 2  # rows written


def main():
    print(f"Reading {INPUT_CSV} ...", flush=True)
    with open(str(INPUT_CSV), encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    print(f"  {len(rows):,} records loaded", flush=True)

    print("Building lookup table ...", flush=True)
    lookup = build_lookup(rows)

    print("Building Excel workbook ...", flush=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    total_rows = 0
    for sheet_name, qcodes in SECTIONS.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit
        n = write_section_sheet(ws, qcodes, lookup)
        total_rows += n
        print(f"  Sheet '{sheet_name}': {n} rows", flush=True)

    print(f"Saving {OUTPUT_XL} ...", flush=True)
    wb.save(str(OUTPUT_XL))

    import os
    size_mb = os.path.getsize(str(OUTPUT_XL)) / 1024 / 1024
    print(f"\nDone. {OUTPUT_XL}")
    print(f"  {total_rows:,} data rows across {len(SECTIONS)} sheets")
    print(f"  File size: {size_mb:.1f} MB")


if __name__ == "__main__":
    main()
